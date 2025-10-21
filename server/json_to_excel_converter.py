#!/usr/bin/env python3
"""
Convert comparison_results_new.json to Excel format with summary details
"""

import json
import pandas as pd
from datetime import datetime
import os

def load_json_data(file_path):
    """Load JSON data from file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except FileNotFoundError:
        print(f"Error: File {file_path} not found")
        return None
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format - {e}")
        return None

def extract_data_for_excel(data):
    """Extract required fields from JSON data for Excel export"""
    results = data.get('results', [])
    
    excel_data = []
    
    for i, result in enumerate(results, 1):
        # Extract basic information
        question = result.get('Question', '')
        answer = result.get('Answer', '')
        old_agentic_answer = result.get('old_Agentic_answer', '')
        new_agentic_answer = result.get('new_Agentic_answer', '')
        
        # Extract scores
        old_score = result.get('old_score', {})
        new_score = result.get('new_score', {})
        
        # Get overall scores from the score objects
        old_overall_score = old_score.get('Overall_Score', 0)
        new_overall_score = new_score.get('Overall_Score', 0)
        
        # User information
        user_data = result.get('User_data', {})
        user_name = user_data.get('User_name', '')
        user_id = user_data.get('UserID', '')
        
        # Question metadata
        question_weights = result.get('Question_weights', {})
        question_complexity = question_weights.get('Question_Complexity', '')
        question_format = question_weights.get('Question_format', '')
        
        # Comparison result
        comparison = result.get('comparison', {})
        better_answer = comparison.get('better_answer', '')
        
        excel_data.append({
            'Row_Number': i,
            'User_Name': user_name,
            'User_ID': user_id,
            'Question_Complexity': question_complexity,
            'Question_Format': question_format,
            'Question': question,
            'Ground_Truth_Answer': answer,
            'Old_Agentic_Answer': old_agentic_answer,
            'New_Agentic_Answer': new_agentic_answer,
            'Old_Overall_Score': old_overall_score,
            'New_Overall_Score': new_overall_score,
            'Better_Answer': better_answer,
            'Score_Difference': new_overall_score - old_overall_score
        })
    
    return excel_data

def create_summary_data(data):
    """Create summary statistics for the bottom of the Excel file"""
    summary_stats = data.get('summary_statistics', {})
    
    # Average scores
    avg_scores = summary_stats.get('average_scores', {})
    old_avg = avg_scores.get('oldOverall', 0)
    new_avg = avg_scores.get('newOverall', 0)
    
    # Comparison results
    comparison_results = summary_stats.get('comparison_results', {})
    old_wins = comparison_results.get('oldWins', 0)
    new_wins = comparison_results.get('newWins', 0)
    ties = comparison_results.get('ties', 0)
    
    # Performance insights
    performance_insights = summary_stats.get('performance_insights', {})
    overall_improvement = performance_insights.get('overall_improvement_percentage', 0)
    
    # Metadata
    metadata = data.get('metadata', {})
    total_queries = metadata.get('total_queries', 0)
    generated_at = metadata.get('generated_at', '')
    
    summary_data = [
        {'Metric': 'Total Queries', 'Value': total_queries, 'Description': 'Total number of questions analyzed'},
        {'Metric': 'Generated At', 'Value': generated_at, 'Description': 'Analysis generation timestamp'},
        {'Metric': '', 'Value': '', 'Description': ''},  # Empty row for spacing
        {'Metric': 'PERFORMANCE COMPARISON', 'Value': '', 'Description': ''},
        {'Metric': 'Old System Average Score', 'Value': f"{old_avg:.3f}", 'Description': 'Average overall score for old system'},
        {'Metric': 'New System Average Score', 'Value': f"{new_avg:.3f}", 'Description': 'Average overall score for new system'},
        {'Metric': 'Overall Improvement', 'Value': f"{overall_improvement:.2f}%", 'Description': 'Percentage improvement from old to new'},
        {'Metric': '', 'Value': '', 'Description': ''},  # Empty row for spacing
        {'Metric': 'WIN/LOSS ANALYSIS', 'Value': '', 'Description': ''},
        {'Metric': 'Old System Wins', 'Value': old_wins, 'Description': 'Number of queries where old system performed better'},
        {'Metric': 'New System Wins', 'Value': new_wins, 'Description': 'Number of queries where new system performed better'},
        {'Metric': 'Ties', 'Value': ties, 'Description': 'Number of queries with equal performance'},
        {'Metric': '', 'Value': '', 'Description': ''},  # Empty row for spacing
        {'Metric': 'DETAILED METRICS', 'Value': '', 'Description': ''},
        {'Metric': 'Old Factuality Score', 'Value': f"{avg_scores.get('oldFactuality', 0):.3f}", 'Description': 'Average factuality score for old system'},
        {'Metric': 'New Factuality Score', 'Value': f"{avg_scores.get('newFactuality', 0):.3f}", 'Description': 'Average factuality score for new system'},
        {'Metric': 'Old Completeness Score', 'Value': f"{avg_scores.get('oldCompleteness', 0):.3f}", 'Description': 'Average completeness score for old system'},
        {'Metric': 'New Completeness Score', 'Value': f"{avg_scores.get('newCompleteness', 0):.3f}", 'Description': 'Average completeness score for new system'},
    ]
    
    return summary_data

def create_excel_file(excel_data, summary_data, output_file):
    """Create Excel file with data and summary"""
    
    # Create main dataframe
    df_main = pd.DataFrame(excel_data)
    
    # Create summary dataframe
    df_summary = pd.DataFrame(summary_data)
    
    # Write to Excel with multiple formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write main data
        df_main.to_excel(writer, sheet_name='Comparison Results', index=False, startrow=0)
        
        # Calculate where to start summary (after main data + some spacing)
        summary_start_row = len(df_main) + 3
        
        # Write summary data
        df_summary.to_excel(writer, sheet_name='Comparison Results', index=False, 
                           startrow=summary_start_row, startcol=0)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Comparison Results']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Style main headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style summary headers
        summary_header_row = summary_start_row + 1
        for cell in worksheet[summary_header_row]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders and formatting to numerical columns
        from openpyxl.styles import NamedStyle, Border, Side
        
        # Create a style for score columns
        score_style = NamedStyle(name="score_style")
        score_style.number_format = '0.000'
        score_style.alignment = Alignment(horizontal="center")
        
        # Apply formatting to score columns
        score_columns = ['J', 'K', 'N']  # Old_Overall_Score, New_Overall_Score, Score_Difference
        for col in score_columns:
            for row in range(2, len(df_main) + 2):  # Skip header row
                cell = worksheet[f'{col}{row}']
                cell.style = score_style
        
        print(f"Excel file created successfully: {output_file}")
        print(f"Main data rows: {len(df_main)}")
        print(f"Summary starting at row: {summary_start_row + 1}")

def main():
    """Main function"""
    # File paths
    input_file = '/Users/telkar.varasree/Downloads/xyne/server/comparison_results_new.json'
    output_file = '/Users/telkar.varasree/Downloads/xyne/server/comparison_analysis_results.xlsx'
    
    print("Loading JSON data...")
    data = load_json_data(input_file)
    
    if data is None:
        print("Failed to load data. Exiting.")
        return
    
    print("Extracting data for Excel...")
    excel_data = extract_data_for_excel(data)
    
    print("Creating summary statistics...")
    summary_data = create_summary_data(data)
    
    print("Creating Excel file...")
    create_excel_file(excel_data, summary_data, output_file)
    
    print(f"\n✅ Conversion completed successfully!")
    print(f"📊 Excel file saved as: {output_file}")
    print(f"📈 Total queries processed: {len(excel_data)}")
    
    # Print some basic statistics
    if excel_data:
        old_scores = [row['Old_Overall_Score'] for row in excel_data]
        new_scores = [row['New_Overall_Score'] for row in excel_data]
        
        print(f"\n📊 Quick Statistics:")
        print(f"   Old System Average: {sum(old_scores)/len(old_scores):.3f}")
        print(f"   New System Average: {sum(new_scores)/len(new_scores):.3f}")
        print(f"   Improvement: {((sum(new_scores)/len(new_scores)) - (sum(old_scores)/len(old_scores))):.3f}")

if __name__ == "__main__":
    main()